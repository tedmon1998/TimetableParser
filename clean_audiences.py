import os
import re
import json
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

def load_audiences():
    """Загружает список валидных аудиторий из info/aud.json"""
    try:
        with open('info/aud.json', 'r', encoding='utf-8') as f:
            audiences = json.load(f)
        return set(audiences)
    except Exception as e:
        print(f"Ошибка при загрузке аудиторий: {e}")
        return set()

def extract_audiences_from_text(text, valid_audiences):
    """Извлекает аудитории из текста дисциплины"""
    if not text:
        return []
    
    text = str(text)
    found_audiences = []
    
    # Ищем все возможные аудитории в тексте
    # Паттерны: А539, У708, К506, Г201, СОКБ, ЭОиДОТ и т.д.
    # Проверяем каждую валидную аудиторию
    for aud in valid_audiences:
        # Ищем точное совпадение (с учетом регистра и границ слова)
        # Используем регулярное выражение для поиска
        pattern = r'\b' + re.escape(aud) + r'\b'
        if re.search(pattern, text, re.IGNORECASE):
            found_audiences.append(aud)
    
    # Если не нашли по списку, пытаемся найти по паттернам
    if not found_audiences:
        # Паттерны для аудиторий: буква + цифры (А539, У708, К506)
        patterns = [
            r'\b([А-ЯЁ][А-ЯЁ]?\d{2,4})\b',  # А539, У708, К506
            r'\b(СОКБ|СОКЦОМиД|ЭОиДОТ|ЭБЦ|ЦАС|УЦ)\b',  # Специальные аудитории
            r'\b(бассейн|зал\s+2|зал\s+гимн)\b',  # Специальные залы
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                # Проверяем, что это валидная аудитория
                for aud in valid_audiences:
                    if aud.upper() == match.upper():
                        if aud not in found_audiences:
                            found_audiences.append(aud)
                        break
    
    return found_audiences

def parse_week_division(text):
    """Определяет разделение по числителю/знаменателю (// или /)"""
    if not text:
        return None, None
    
    text = str(text)
    
    # Ищем разделитель "//" (числитель/знаменатель)
    if '//' in text:
        parts = text.split('//', 1)
        numerator = parts[0].strip()
        denominator = parts[1].strip() if len(parts) > 1 else ''
        return numerator, denominator
    
    # Ищем разделитель "/" (может быть числитель/знаменатель или просто разделитель)
    if '/' in text:
        # Проверяем, не является ли это частью аудитории (например, "п/г")
        if 'п/г' in text.lower() or 'подгруппа' in text.lower():
            return None, None
        
        parts = text.split('/', 1)
        numerator = parts[0].strip()
        denominator = parts[1].strip() if len(parts) > 1 else ''
        # Если обе части не пустые, это может быть числитель/знаменатель
        if numerator and denominator:
            return numerator, denominator
    
    return None, None

def extract_lecture_type(text):
    """Определяет тип занятия (лекция/практика)"""
    if not text:
        return None
    
    text = str(text).lower()
    
    # Проверяем наличие подгрупп - если есть, то это практика
    if 'п/г' in text or 'подгруппа' in text:
        return 'практика'
    
    # Проверяем маркеры типа занятия
    if '(лек)' in text or 'лек' in text or 'лекция' in text:
        return 'лекция'
    
    if '(пр)' in text or 'практика' in text:
        return 'практика'
    
    # Если есть разделение по подгруппам (п/г 1, п/г 2), это практика
    if re.search(r'п/г\s*\d+', text, re.IGNORECASE):
        return 'практика'
    
    return None

def split_multiple_disciplines(text, valid_audiences):
    """Разбивает текст на несколько дисциплин, если они есть"""
    if not text:
        return [text]
    
    text = str(text)
    disciplines = []
    
    # Ищем разделители между дисциплинами
    # Паттерн: валидная аудитория + пробел + заглавная буква (начало нового предмета)
    # Или: валидная аудитория + пробел + "//" (начало знаменателя)
    
    # Создаем паттерн из всех валидных аудиторий
    aud_pattern = '|'.join([re.escape(aud) for aud in sorted(valid_audiences, key=len, reverse=True)])
    
    # Паттерн: аудитория + пробел + заглавная буква или "//"
    pattern = rf'\b({aud_pattern})\s+([А-ЯЁ]|//)'
    
    matches = list(re.finditer(pattern, text, re.IGNORECASE))
    
    if not matches:
        # Если не нашли разделители, возвращаем весь текст как одну дисциплину
        return [text]
    
    # Разбиваем текст по найденным разделителям
    last_pos = 0
    for i, match in enumerate(matches):
        # Берем текст до разделителя (включая аудиторию)
        # Разделитель - это позиция перед началом нового предмета
        split_pos = match.start() + len(match.group(1))  # Позиция после аудитории
        
        part = text[last_pos:split_pos].strip()
        if part:
            disciplines.append(part)
        
        last_pos = split_pos
    
    # Добавляем оставшуюся часть
    if last_pos < len(text):
        remaining = text[last_pos:].strip()
        if remaining:
            disciplines.append(remaining)
    
    # Если получили только одну часть, возвращаем исходный текст
    if len(disciplines) <= 1:
        return [text]
    
    return disciplines

def clean_discipline_name(text, audience_to_remove=None):
    """Очищает название дисциплины от аудитории, оставляя только название и подгруппу"""
    if not text:
        return text
    
    text = str(text)
    
    # Если указана аудитория для удаления, убираем её
    if audience_to_remove:
        # Убираем аудиторию из текста
        pattern = r'\b' + re.escape(audience_to_remove) + r'\b'
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    # Убираем все валидные аудитории из текста (на случай, если остались другие)
    # Но только если они не являются частью названия дисциплины
    # Сначала убираем паттерны типа ", А539", ", У708" и т.д.
    text = re.sub(r',\s*([А-ЯЁ][А-ЯЁ]?\d{2,4}|СОКБ|СОКЦОМиД|ЭОиДОТ|ЭБЦ|ЦАС|УЦ|бассейн|зал\s+2|зал\s+гимн)\b', '', text, flags=re.IGNORECASE)
    
    # Убираем разделители "//" в середине текста (но не в начале/конце, они важны)
    text = re.sub(r'\s*//\s*', ' ', text)
    
    # Убираем лишние пробелы и запятые
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r',\s*,', ',', text)  # Двойные запятые
    text = re.sub(r',\s*$', '', text)  # Запятая в конце
    text = text.strip(', ').strip()
    
    return text

def clean_subject_name_final(text, valid_audiences=None):
    """Очищает название предмета от лишних символов: (лек), (пр), п/г, аудитории и т.д."""
    if not text:
        return text
    
    text = str(text)
    
    # Убираем "(лек)", "(пр)" и подобные
    text = re.sub(r'\(лек\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(пр\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(лек\s*\d+\s*ч\)', '', text, flags=re.IGNORECASE)  # (лек 8 ч)
    text = re.sub(r'\(практика\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(лекция\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(лекция\s*\d+\s*ч\)', '', text, flags=re.IGNORECASE)  # (лекция 8 ч)
    text = re.sub(r'\(лекция\s+\d+\s+ч\)', '', text, flags=re.IGNORECASE)  # (лекция 8 ч) с пробелами
    text = re.sub(r'\(24\s*ч\)', '', text, flags=re.IGNORECASE)  # (24 ч)
    
    # Убираем подгруппы: п/г 1, п/г 2, п/г1, п/г2 и т.д. (более агрессивно)
    # Сначала убираем с запятой, потом любые вхождения (включая без пробела после запятой)
    text = re.sub(r',\s*п/г\s*\d+', '', text, flags=re.IGNORECASE)
    text = re.sub(r',п/г\s*\d+', '', text, flags=re.IGNORECASE)  # Без пробела после запятой
    text = re.sub(r'\s+п/г\s*\d+', '', text, flags=re.IGNORECASE)  # С пробелом перед
    text = re.sub(r'п/г\s*\d+', '', text, flags=re.IGNORECASE)  # Любое вхождение п/г
    text = re.sub(r',\s*подгруппа\s*\d+', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s+подгруппа\s*\d+', '', text, flags=re.IGNORECASE)
    text = re.sub(r'подгруппа\s*\d+', '', text, flags=re.IGNORECASE)  # Любое вхождение подгруппа
    
    # Убираем все валидные аудитории из текста
    if valid_audiences:
        for aud in valid_audiences:
            # Убираем аудиторию в разных форматах: ", А539", " А539", "А539"
            pattern = r'[, ]\s*' + re.escape(aud) + r'\b'
            text = re.sub(pattern, '', text, flags=re.IGNORECASE)
            pattern = r'\b' + re.escape(aud) + r'\b'
            text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    # Убираем разделители "//" в середине текста
    text = re.sub(r'\s*//\s*', ' ', text)
    
    # Убираем лишние пробелы и запятые
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r',\s*,', ',', text)  # Двойные запятые
    text = re.sub(r',\s*$', '', text)  # Запятая в конце
    text = text.strip(', ').strip()
    
    return text

def split_teachers(teacher_text):
    """Разделяет преподавателей по '/' - первый для числителя, второй для знаменателя"""
    if not teacher_text:
        return None, None
    
    teacher_text = str(teacher_text).strip()
    if '/' in teacher_text:
        parts = teacher_text.split('/', 1)  # Разделяем только по первому '/'
        numerator_teacher = parts[0].strip() if parts[0] else None
        denominator_teacher = parts[1].strip() if len(parts) > 1 and parts[1] else None
        return numerator_teacher, denominator_teacher
    
    # Если нет '/', то один преподаватель для обеих недель
    return teacher_text, teacher_text

def split_teachers_by_subgroups(teacher_text):
    """Разделяет преподавателей по ';' - каждый для своей подгруппы"""
    if not teacher_text:
        return []
    
    teacher_text = str(teacher_text).strip()
    if ';' in teacher_text:
        teachers = [t.strip() for t in teacher_text.split(';') if t.strip()]
        return teachers
    
    # Если нет ';', то один преподаватель для всех подгрупп
    return [teacher_text] if teacher_text else []

def extract_subgroups_from_text(text):
    """Извлекает номера подгрупп из текста (п/г 1, п/г 2 и т.д.)"""
    if not text:
        return []
    
    text = str(text)
    subgroups = []
    
    # Ищем паттерны: п/г 1, п/г 2, п/г1, п/г2 и т.д.
    pattern = r'п/г\s*(\d+)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    for match in matches:
        subgroup_num = int(match)
        if subgroup_num not in subgroups:
            subgroups.append(subgroup_num)
    
    # Ищем паттерны: подгруппа 1, подгруппа 2 и т.д.
    pattern = r'подгруппа\s*(\d+)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    for match in matches:
        subgroup_num = int(match)
        if subgroup_num not in subgroups:
            subgroups.append(subgroup_num)
    
    return sorted(subgroups)

def extract_audience_for_subgroup(text, subgroup_num, valid_audiences):
    """Извлекает аудиторию для конкретной подгруппы из текста"""
    if not text or not subgroup_num:
        return None
    
    text = str(text)
    # Ищем паттерн: п/г N, АУДИТОРИЯ или п/г N АУДИТОРИЯ
    # Ищем участок текста, связанный с этой подгруппой
    pattern = rf'п/г\s*{subgroup_num}[,\s]+([А-ЯЁ][А-ЯЁ]?\d{{2,4}}|СОКБ|СОКЦОМиД|ЭОиДОТ|ЭБЦ|ЦАС|УЦ|бассейн|зал\s+2|зал\s+гимн)'
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        potential_aud = match.group(1)
        # Проверяем, что это валидная аудитория
        if potential_aud in valid_audiences:
            return potential_aud
    
    # Альтернативный паттерн: подгруппа N, АУДИТОРИЯ
    pattern = rf'подгруппа\s*{subgroup_num}[,\s]+([А-ЯЁ][А-ЯЁ]?\d{{2,4}}|СОКБ|СОКЦОМиД|ЭОиДОТ|ЭБЦ|ЦАС|УЦ|бассейн|зал\s+2|зал\s+гимн)'
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        potential_aud = match.group(1)
        if potential_aud in valid_audiences:
            return potential_aud
    
    return None

def process_discipline_text(text, valid_audiences, teacher_text=None):
    """Обрабатывает текст дисциплины и возвращает список записей (по одной на каждую дисциплину/неделю)"""
    if not text:
        return []
    
    text = str(text)
    results = []
    
    # Проверяем, есть ли разделение преподавателей по ';' (для подгрупп)
    teachers_by_subgroups = split_teachers_by_subgroups(teacher_text) if teacher_text else []
    subgroups_in_text = extract_subgroups_from_text(text)
    
    # Если есть разделение по ';' и есть подгруппы в тексте, обрабатываем отдельно
    if len(teachers_by_subgroups) > 1 and len(subgroups_in_text) > 0:
        # Определяем тип занятия
        lecture_type = extract_lecture_type(text)
        if not lecture_type:
            lecture_type = 'практика'  # Если есть подгруппы, это практика
        
        # Очищаем название дисциплины
        clean_name = clean_subject_name_final(text, valid_audiences)
        
        # Сопоставляем преподавателей с подгруппами
        for i, subgroup_num in enumerate(subgroups_in_text):
            teacher_for_subgroup = teachers_by_subgroups[i] if i < len(teachers_by_subgroups) else teachers_by_subgroups[-1]
            
            # Извлекаем аудиторию для этой конкретной подгруппы
            aud_for_subgroup = extract_audience_for_subgroup(text, subgroup_num, valid_audiences)
            
            # Если не нашли специфичную аудиторию, берем все аудитории из текста
            if not aud_for_subgroup:
                all_audiences = extract_audiences_from_text(text, valid_audiences)
                aud_for_subgroup = all_audiences[0] if all_audiences else ''
            
            result_entry = {
                'audience': aud_for_subgroup,
                'subject_name': clean_name,
                'lecture_type': lecture_type or '',
                'teacher': teacher_for_subgroup,
                'subgroup': subgroup_num
            }
            results.append(result_entry)
        
        # Если создали записи для подгрупп, возвращаем их
        if results:
            return results
    
    # Иначе обрабатываем как обычно (разделение по '/' для числителя/знаменателя)
    # Разделяем преподавателей, если они есть
    numerator_teacher, denominator_teacher = split_teachers(teacher_text) if teacher_text else (None, None)
    
    # Определяем тип занятия для всего текста
    lecture_type = extract_lecture_type(text)
    if not lecture_type and ('п/г' in text.lower() or 'подгруппа' in text.lower()):
        lecture_type = 'практика'
    
    # Проверяем, есть ли разделение по неделям (//) во всем тексте
    if '//' in text:
        # Есть разделение по числителю/знаменателю
        # Разбиваем на части по "//"
        parts = text.split('//')
        
        # Числитель - все части до последнего "//"
        numerator_parts = parts[:-1] if len(parts) > 1 else []
        numerator_text = ' '.join(numerator_parts).strip()
        
        # Знаменатель - последняя часть после "//"
        denominator_text = parts[-1].strip() if len(parts) > 1 else ''
        
        # Извлекаем аудитории из числителя (убираем дубликаты)
        numerator_audiences_unique = []
        if numerator_text:
            numerator_audiences = extract_audiences_from_text(numerator_text, valid_audiences)
            # Убираем дубликаты, сохраняя порядок
            for aud in numerator_audiences:
                if aud not in numerator_audiences_unique:
                    numerator_audiences_unique.append(aud)
        
        # Извлекаем аудитории из знаменателя (убираем дубликаты)
        denominator_audiences_unique = []
        if denominator_text:
            denominator_audiences = extract_audiences_from_text(denominator_text, valid_audiences)
            # Убираем дубликаты, сохраняя порядок
            for aud in denominator_audiences:
                if aud not in denominator_audiences_unique:
                    denominator_audiences_unique.append(aud)
        
        # Создаем отдельные строки для числителя (полный текст числителя, включая аудиторию)
        for aud in numerator_audiences_unique:
            # Очищаем название от лишних символов
            clean_name = clean_subject_name_final(numerator_text, valid_audiences)
            result_entry = {
                'audience': aud,
                'subject_name': clean_name,
                'lecture_type': lecture_type or ''
            }
            # Добавляем преподавателя для числителя
            if numerator_teacher:
                result_entry['teacher'] = numerator_teacher
                result_entry['week_type'] = 'числитель'
            results.append(result_entry)
        
        # Создаем отдельные строки для знаменателя (полный текст знаменателя, включая аудиторию)
        for aud in denominator_audiences_unique:
            # Очищаем название от лишних символов
            clean_name = clean_subject_name_final(denominator_text, valid_audiences)
            result_entry = {
                'audience': aud,
                'subject_name': clean_name,
                'lecture_type': lecture_type or ''
            }
            # Добавляем преподавателя для знаменателя
            if denominator_teacher:
                result_entry['teacher'] = denominator_teacher
                result_entry['week_type'] = 'знаменатель'
            results.append(result_entry)
    else:
        # Нет разделения по неделям - разбиваем на несколько дисциплин
        disciplines = split_multiple_disciplines(text, valid_audiences)
        
        # Обрабатываем каждую дисциплину отдельно (обе недели)
        for disc_text in disciplines:
            audiences = extract_audiences_from_text(disc_text, valid_audiences)
            # Убираем дубликаты, сохраняя порядок
            audiences_unique = []
            for aud in audiences:
                if aud not in audiences_unique:
                    audiences_unique.append(aud)
            
            # Создаем одну строку для каждой уникальной аудитории
            for aud in audiences_unique:
                # Очищаем название дисциплины от аудитории, оставляя только название и подгруппу
                clean_name = clean_discipline_name(disc_text, aud)
                # Финальная очистка от (лек), (пр), п/г и т.д.
                clean_name = clean_subject_name_final(clean_name, valid_audiences)
                result_entry = {
                    'audience': aud,
                    'subject_name': clean_name,
                    'lecture_type': lecture_type or ''
                }
                # Если есть преподаватель (без разделения), добавляем его для обеих недель
                if numerator_teacher and numerator_teacher == denominator_teacher:
                    result_entry['teacher'] = numerator_teacher
                    result_entry['week_type'] = 'обе недели'
                results.append(result_entry)
    
    # Если не нашли ни одной аудитории, создаем запись без аудитории
    if not results:
        # Финальная очистка от (лек), (пр), п/г и т.д.
        clean_name = clean_subject_name_final(text, valid_audiences)
        result_entry = {
            'audience': '',
            'subject_name': clean_name,
            'lecture_type': lecture_type or ''
        }
        # Если есть преподаватель (без разделения), добавляем его
        if numerator_teacher and numerator_teacher == denominator_teacher:
            result_entry['teacher'] = numerator_teacher
            result_entry['week_type'] = 'обе недели'
        results.append(result_entry)
    
    return results

def process_csv_file(input_file, output_file, valid_audiences):
    """Обрабатывает CSV файл и создает очищенную версию"""
    results = []
    
    with open(input_file, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        
        for row in reader:
            subject_name = row.get('subject_name', '')
            teacher_fio = row.get('fio', '') or row.get('teacher', '')
            
            # Обрабатываем текст дисциплины - получаем список записей
            processed_list = process_discipline_text(subject_name, valid_audiences, teacher_fio)
            
            # Создаем отдельную запись для каждой дисциплины/аудитории
            for processed in processed_list:
                new_row = row.copy()
                new_row['audience'] = processed['audience']
                new_row['subject_name'] = processed['subject_name']
                new_row['lecture_type'] = processed['lecture_type']
                # Добавляем преподавателя и тип недели, если они есть
                if 'teacher' in processed:
                    new_row['fio'] = processed['teacher']
                    if 'teacher' in new_row:
                        new_row['teacher'] = processed['teacher']
                if 'week_type' in processed:
                    new_row['week_type'] = processed['week_type']
                elif 'week' in new_row and not new_row.get('week_type'):
                    # Если week_type не был создан, используем значение из исходного поля week
                    new_row['week_type'] = new_row.get('week', '')
                if 'subgroup' in processed:
                    new_row['subgroup'] = processed['subgroup']
                # Убеждаемся, что group_name есть, даже если оно называется group
                if 'group' in new_row and not new_row.get('group_name'):
                    new_row['group_name'] = new_row.get('group', '')
                # Удаляем старые поля, если они есть
                new_row.pop('audience_numerator', None)
                new_row.pop('audience_denominator', None)
                
                results.append(new_row)
    
    # Определяем новые заголовки
    new_fieldnames = []
    # Порядок: day_of_week, pair_number, subject_name, lecture_type, audience
    priority_fields = ['day_of_week', 'pair_number', 'subject_name', 'lecture_type', 'audience']
    
    # Добавляем приоритетные поля
    for field in priority_fields:
        if field in fieldnames:
            new_fieldnames.append(field)
    
    # Добавляем остальные поля (кроме старых audience_numerator, audience_denominator)
    for field in fieldnames:
        if field not in new_fieldnames and field not in ['audience_numerator', 'audience_denominator']:
            new_fieldnames.append(field)
    
    # Убеждаемся, что audience есть в заголовках
    if 'audience' not in new_fieldnames:
        # Вставляем после subject_name или в конец
        if 'subject_name' in new_fieldnames:
            idx = new_fieldnames.index('subject_name') + 1
            new_fieldnames.insert(idx, 'audience')
        else:
            new_fieldnames.append('audience')
    
    # Убеждаемся, что lecture_type есть в заголовках
    if 'lecture_type' not in new_fieldnames:
        # Вставляем после subject_name или в конец
        if 'subject_name' in new_fieldnames:
            idx = new_fieldnames.index('subject_name') + 1
            new_fieldnames.insert(idx, 'lecture_type')
        else:
            new_fieldnames.append('lecture_type')
    
    # Убеждаемся, что fio, week_type, subgroup и group_name есть в заголовках, если они используются
    if 'fio' not in new_fieldnames:
        # Проверяем, используется ли fio в результатах
        for result in results:
            if 'fio' in result and result.get('fio'):
                new_fieldnames.append('fio')
                break
    if 'teacher' not in new_fieldnames:
        # Проверяем, используется ли teacher в результатах
        for result in results:
            if 'teacher' in result and result.get('teacher'):
                new_fieldnames.append('teacher')
                break
    if 'week_type' not in new_fieldnames:
        # Проверяем, используется ли week_type в результатах
        for result in results:
            if 'week_type' in result and result.get('week_type'):
                new_fieldnames.append('week_type')
                break
    if 'group_name' not in new_fieldnames:
        # Проверяем, используется ли group_name в результатах
        for result in results:
            if 'group_name' in result and result.get('group_name'):
                new_fieldnames.append('group_name')
                break
        # Если group_name не найден, но есть group, добавляем group_name
        if 'group_name' not in new_fieldnames:
            for result in results:
                if 'group' in result and result.get('group'):
                    new_fieldnames.append('group_name')
                    break
    if 'subgroup' not in new_fieldnames:
        # Проверяем, используется ли subgroup в результатах
        for result in results:
            if 'subgroup' in result and result.get('subgroup'):
                new_fieldnames.append('subgroup')
                break
    
    # Убеждаемся, что все важные поля присутствуют в заголовках
    important_fields = ['fio', 'teacher', 'group_name', 'week_type']
    for field in important_fields:
        if field not in new_fieldnames:
            # Проверяем, есть ли это поле хотя бы в одной записи
            for result in results:
                if field in result:
                    new_fieldnames.append(field)
                    break
    
    # Логируем финальные заголовки для отладки
    print(f"DEBUG: Финальные заголовки CSV: {new_fieldnames}")
    if results:
        print(f"DEBUG: Первая запись результата: {list(results[0].keys())}")
        print(f"DEBUG: Значения первой записи - fio: {results[0].get('fio')}, teacher: {results[0].get('teacher')}, group_name: {results[0].get('group_name')}, week_type: {results[0].get('week_type')}")
    
    # Сохраняем результат
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=new_fieldnames)
        writer.writeheader()
        writer.writerows(results)
    
    return len(results)

def process_excel_file(input_file, output_file, valid_audiences):
    """Обрабатывает Excel файл и создает очищенную версию"""
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active
    
    # Читаем заголовки
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            headers.append(str(cell.value))
        else:
            headers.append(f'Column{col}')
    
    # Находим индексы нужных колонок
    subject_idx = None
    for i, header in enumerate(headers):
        if 'subject_name' in header.lower() or 'дисциплина' in header.lower():
            subject_idx = i
            break
    
    if subject_idx is None:
        print("Не найдена колонка с дисциплиной!")
        return 0
    
    # Добавляем новые колонки после subject_name
    new_headers = headers.copy()
    insert_idx = subject_idx + 1
    if 'audience_numerator' not in new_headers:
        new_headers.insert(insert_idx, 'audience_numerator')
        insert_idx += 1
    if 'audience_denominator' not in new_headers:
        new_headers.insert(insert_idx, 'audience_denominator')
        insert_idx += 1
    if 'lecture_type' not in new_headers:
        new_headers.insert(insert_idx, 'lecture_type')
    
    # Обрабатываем данные
    results = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_idx)
            row_data[header] = cell.value if cell.value else ''
        
        subject_name = row_data.get(headers[subject_idx], '')
        
        # Находим колонку с преподавателем
        teacher_idx = None
        for i, header in enumerate(headers):
            if 'fio' in header.lower() or 'teacher' in header.lower() or 'преподаватель' in header.lower():
                teacher_idx = i
                break
        
        teacher_fio = row_data.get(headers[teacher_idx], '') if teacher_idx is not None else ''
        
        # Обрабатываем текст дисциплины - получаем список записей
        processed_list = process_discipline_text(subject_name, valid_audiences, teacher_fio)
        
        # Создаем отдельную запись для каждой дисциплины/аудитории
        for processed in processed_list:
            new_row = row_data.copy()
            new_row['audience'] = processed['audience']
            new_row['subject_name'] = processed['subject_name']
            new_row['lecture_type'] = processed['lecture_type']
            # Добавляем преподавателя и тип недели, если они есть
            if 'teacher' in processed:
                # Обновляем существующую колонку fio или teacher
                if teacher_idx is not None:
                    new_row[headers[teacher_idx]] = processed['teacher']
                elif 'fio' in new_row:
                    new_row['fio'] = processed['teacher']
            if 'week_type' in processed:
                new_row['week_type'] = processed['week_type']
            if 'subgroup' in processed:
                new_row['subgroup'] = processed['subgroup']
            # Удаляем старые поля, если они есть
            new_row.pop('audience_numerator', None)
            new_row.pop('audience_denominator', None)
            
            results.append(new_row)
    
    # Определяем новые заголовки (аналогично CSV)
    new_headers = []
    priority_fields = ['day_of_week', 'pair_number', 'subject_name', 'lecture_type', 'audience']
    
    # Добавляем приоритетные поля
    for field in priority_fields:
        if field in headers:
            new_headers.append(field)
    
    # Добавляем остальные поля (кроме старых audience_numerator, audience_denominator)
    for field in headers:
        if field not in new_headers and field not in ['audience_numerator', 'audience_denominator']:
            new_headers.append(field)
    
    # Убеждаемся, что audience и lecture_type есть в заголовках
    if 'audience' not in new_headers:
        if 'subject_name' in new_headers:
            idx = new_headers.index('subject_name') + 1
            new_headers.insert(idx, 'audience')
        else:
            new_headers.append('audience')
    
    if 'lecture_type' not in new_headers:
        if 'subject_name' in new_headers:
            idx = new_headers.index('subject_name') + 1
            new_headers.insert(idx, 'lecture_type')
        else:
            new_headers.append('lecture_type')
    
    # Убеждаемся, что week_type и subgroup есть в заголовках, если они используются
    if 'week_type' not in new_headers:
        for result in results:
            if 'week_type' in result and result.get('week_type'):
                new_headers.append('week_type')
                break
    if 'subgroup' not in new_headers:
        for result in results:
            if 'subgroup' in result and result.get('subgroup'):
                new_headers.append('subgroup')
                break
    
    # Создаем новый файл
    wb_new = Workbook()
    ws_new = wb_new.active
    
    # Записываем заголовки
    for col_idx, header in enumerate(new_headers, 1):
        cell = ws_new.cell(1, col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Записываем данные
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, header in enumerate(new_headers, 1):
            value = row_data.get(header, '')
            ws_new.cell(row_idx, col_idx, value=value)
    
    # Автоподбор ширины столбцов
    for col in ws_new.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if max_length < length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_new.column_dimensions[col_letter].width = adjusted_width
    
    # Замораживаем первую строку
    ws_new.freeze_panes = 'A2'
    
    wb_new.save(output_file)
    return len(results)

def save_to_database(csv_file):
    """Сохраняет данные из CSV файла в PostgreSQL базу данных"""
    try:
        import psycopg2
        from psycopg2.extras import execute_values
    except ImportError:
        print("Ошибка: библиотека psycopg2 не установлена. Установите её командой: pip install psycopg2-binary")
        return False
    
    # Параметры подключения к БД
    db_config = {
        'host': 'edro.su',
        'port': 50003,
        'user': 'edro',
        'password': 'Pg123!',
        'database': 'test_sursu_timetable'
    }
    
    try:
        # Подключаемся к БД
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # Создаем таблицу, если её нет
        create_table_query = """
        CREATE TABLE IF NOT EXISTS timetable_cleaned (
            id SERIAL PRIMARY KEY,
            day_of_week VARCHAR(50),
            pair_number INTEGER,
            subject_name TEXT,
            lecture_type VARCHAR(50),
            audience VARCHAR(50),
            fio TEXT,
            teacher TEXT,
            group_name VARCHAR(50),
            week_type VARCHAR(50),
            subgroup INTEGER,
            institute TEXT,
            course VARCHAR(10),
            direction TEXT,
            department TEXT,
            is_external BOOLEAN,
            is_remote BOOLEAN,
            num_subgroups INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Таблица timetable_cleaned создана или уже существует")
        
        # Очищаем таблицу перед вставкой новых данных
        cursor.execute("TRUNCATE TABLE timetable_cleaned")
        conn.commit()
        print("Таблица timetable_cleaned очищена")
        
        # Читаем данные из CSV файла
        rows_to_insert = []
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            
            # Логируем доступные поля для отладки
            print(f"DEBUG: Поля в CSV файле: {fieldnames}")
            
            for row in reader:
                # Вспомогательная функция для преобразования в int
                def to_int(value):
                    if not value or value == '':
                        return None
                    try:
                        return int(value)
                    except (ValueError, TypeError):
                        return None
                
                # Вспомогательная функция для преобразования в bool
                def to_bool(value):
                    if not value or value == '':
                        return False
                    if isinstance(value, bool):
                        return value
                    if isinstance(value, str):
                        return value.lower() in ('true', '1', 'yes', 't')
                    return bool(value)
                
                # Подготавливаем данные для вставки
                # Получаем week_type из разных возможных полей
                week_type_value = (row.get('week_type', '') or 
                                 row.get('week', '') or 
                                 row.get('week_ru', '') or 
                                 None)
                
                # Получаем group_name из разных возможных полей
                group_name_value = (row.get('group_name', '') or 
                                  row.get('group', '') or 
                                  None)
                
                # Получаем fio и teacher
                fio_value = (row.get('fio', '') or 
                           row.get('teacher', '') or 
                           None)
                teacher_value = (row.get('teacher', '') or 
                               row.get('fio', '') or 
                               None)
                
                # Логируем первую строку для отладки
                if len(rows_to_insert) == 0:
                    print(f"DEBUG: Первая строка из CSV:")
                    print(f"  fio={row.get('fio')}, teacher={row.get('teacher')}")
                    print(f"  group={row.get('group')}, group_name={row.get('group_name')}")
                    print(f"  week={row.get('week')}, week_type={row.get('week_type')}")
                    print(f"  fio_value={fio_value}, teacher_value={teacher_value}")
                    print(f"  group_name_value={group_name_value}, week_type_value={week_type_value}")
                
                values = (
                    row.get('day_of_week', '') or None,
                    to_int(row.get('pair_number', '')),
                    row.get('subject_name', '') or None,
                    row.get('lecture_type', '') or None,
                    row.get('audience', '') or None,
                    fio_value,
                    teacher_value,
                    group_name_value,
                    week_type_value,
                    to_int(row.get('subgroup', '')),
                    row.get('institute', '') or None,
                    row.get('course', '') or None,
                    row.get('direction', '') or None,
                    row.get('department', '') or None,
                    to_bool(row.get('is_external', '')),
                    to_bool(row.get('is_remote', '')),
                    to_int(row.get('num_subgroups', ''))
                )
                rows_to_insert.append(values)
        
        # Вставляем данные пакетами
        if rows_to_insert:
            insert_query = """
            INSERT INTO timetable_cleaned (
                day_of_week, pair_number, subject_name, lecture_type, audience,
                fio, teacher, group_name, week_type, subgroup,
                institute, course, direction, department,
                is_external, is_remote, num_subgroups
            ) VALUES %s
            """
            
            execute_values(cursor, insert_query, rows_to_insert)
            conn.commit()
            
            print(f"Успешно сохранено {len(rows_to_insert)} записей в базу данных")
        
        cursor.close()
        conn.close()
        return True
        
    except psycopg2.Error as e:
        print(f"Ошибка при работе с базой данных: {e}")
        import traceback
        traceback.print_exc()
        return False
    except Exception as e:
        print(f"Ошибка при сохранении в базу данных: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    import glob
    
    # Загружаем валидные аудитории
    valid_audiences = load_audiences()
    print(f"Загружено {len(valid_audiences)} валидных аудиторий")
    
    # Ищем исходные файлы (без _cleaned)
    csv_input = None
    excel_input = None
    
    # Ищем CSV файл
    csv_files = glob.glob('output/timetable/timetable_processed.csv')
    if csv_files:
        csv_input = csv_files[0]
    
    # Ищем Excel файл (исключаем временные файлы)
    excel_files = [f for f in glob.glob('output/timetable/timetable_processed.xlsx') 
                   if not os.path.basename(f).startswith('~$')]
    if excel_files:
        excel_input = excel_files[0]
    
    if not csv_input and not excel_input:
        print("Не найдены файлы timetable_processed.csv или timetable_processed.xlsx в output/timetable/")
        return
    
    # Фиксированные имена выходных файлов
    csv_output = 'output/timetable/timetable_processed_cleaned.csv'
    excel_output = 'output/timetable/timetable_processed_cleaned.xlsx'
    
    # Обрабатываем CSV
    if csv_input:
        print(f"\nОбрабатываем CSV файл: {csv_input}")
        try:
            count = process_csv_file(csv_input, csv_output, valid_audiences)
            print(f"Обработано записей: {count}")
            print(f"Результат сохранен в: {csv_output}")
        except Exception as e:
            print(f"Ошибка при обработке CSV файла: {e}")
            import traceback
            traceback.print_exc()
    
    # Обрабатываем Excel
    if excel_input:
        print(f"\nОбрабатываем Excel файл: {excel_input}")
        try:
            count = process_excel_file(excel_input, excel_output, valid_audiences)
            print(f"Обработано записей: {count}")
            print(f"Результат сохранен в: {excel_output}")
        except Exception as e:
            print(f"Ошибка при обработке Excel файла: {e}")
            import traceback
            traceback.print_exc()
    
    # Сохраняем в базу данных
    # Используем CSV файл, так как он уже обработан
    if os.path.exists(csv_output):
        print(f"\nСохраняем данные в базу данных...")
        try:
            if save_to_database(csv_output):
                print("Данные успешно сохранены в базу данных")
            else:
                print("Не удалось сохранить данные в базу данных")
        except Exception as e:
            print(f"Ошибка при сохранении в базу данных: {e}")
            import traceback
            traceback.print_exc()
    elif os.path.exists(excel_output):
        # Если CSV нет, но есть Excel, создаем временный CSV для БД
        print(f"\nСоздаем временный CSV для сохранения в БД...")
        try:
            # Читаем Excel и создаем CSV
            wb = load_workbook(excel_output, data_only=True)
            ws = wb.active
            
            # Читаем заголовки
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(1, col)
                if cell.value:
                    headers.append(str(cell.value))
            
            # Создаем временный CSV файл
            temp_csv = csv_output.replace('.csv', '_temp_for_db.csv')
            with open(temp_csv, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                
                for row_idx in range(2, ws.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row_idx, col_idx)
                        row_data[header] = cell.value if cell.value else ''
                    writer.writerow(row_data)
            
            if save_to_database(temp_csv):
                print("Данные успешно сохранены в базу данных")
                # Удаляем временный файл
                os.remove(temp_csv)
            else:
                print("Не удалось сохранить данные в базу данных")
        except Exception as e:
            print(f"Ошибка при создании временного CSV или сохранении в БД: {e}")
            import traceback
            traceback.print_exc()

if __name__ == '__main__':
    main()
