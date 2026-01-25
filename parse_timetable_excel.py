import os
import re
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import json

# Импортируем функции из process_timetable.py
from process_timetable import load_teacher_names, normalize_short_fio

# Маппинг дней недели
DAYS_OF_WEEK = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']

# Префиксы аудиторий
AUDIENCE_PREFIXES = ['У', 'К', 'А', 'Г', 'м/зал', 'бассейн', 'п/б', 'УЦ', 'л/б', 'зал 2', 'зал гимн', 
                     'ЭБЦ', 'ЦАС', 'СОКЦОМиД', 'ЭОиДОТ', 'С', 'СОКБ']

def is_audience(text):
    """Проверяет, является ли текст аудиторией"""
    if not text or not isinstance(text, str):
        return False
    text = text.strip()
    for prefix in AUDIENCE_PREFIXES:
        if text.startswith(prefix):
            return True
    return False

def parse_pair_number(pair_str):
    """Парсит номер пары, может быть диапазон типа '1-2'"""
    if not pair_str:
        return []
    pair_str = str(pair_str).strip()
    
    # Если это диапазон типа "1-2"
    if '-' in pair_str:
        try:
            start, end = map(int, pair_str.split('-'))
            return list(range(start, end + 1))
        except:
            pass
    
    # Если это одно число
    try:
        return [int(pair_str)]
    except:
        return []

def parse_week_type(text):
    """Определяет тип недели по наличию '/' или '//'
    Возвращает список типов недель для разных частей текста"""
    if not text or not isinstance(text, str):
        return ['обе недели']
    
    text = text.strip()
    week_types = []
    
    # Если есть "//" - значит одна пара по числителю или знаменателю
    if '//' in text:
        parts = text.split('//')
        if len(parts) == 2:
            # Если до // есть текст - числитель, если после - знаменатель
            if parts[0].strip() and not parts[1].strip():
                return ['числитель']
            elif parts[1].strip() and not parts[0].strip():
                return ['знаменатель']
            elif parts[0].strip() and parts[1].strip():
                # Обе части заполнены - значит разные пары
                return ['числитель', 'знаменатель']
        return ['обе недели']
    
    # Если есть "/" (одиночный) - значит разные пары по числителю и знаменателю
    if '/' in text and '//' not in text:
        parts = text.split('/')
        if len(parts) == 2:
            if parts[0].strip() and parts[1].strip():
                return ['числитель', 'знаменатель']
            elif parts[0].strip():
                return ['числитель']
            elif parts[1].strip():
                return ['знаменатель']
        return ['обе недели']
    
    return ['обе недели']

def parse_lecture_type(text):
    """Определяет тип занятия: лекция или практика"""
    if not text or not isinstance(text, str):
        return 'практика'
    
    text = text.lower()
    if '(лек)' in text or 'лекция' in text:
        return 'лекция'
    elif '(пр)' in text or 'практика' in text:
        return 'практика'
    else:
        # Если есть "лекция" в тексте, но без скобок
        if 'лек' in text:
            return 'лекция'
        return 'практика'

def extract_audience(text):
    """Извлекает аудиторию из текста"""
    if not text or not isinstance(text, str):
        return ''
    
    text = text.strip()
    audiences = []
    
    # Разделяем по запятым и ищем аудитории
    parts = re.split(r'[,;]', text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # Проверяем, является ли часть аудиторией
        for prefix in AUDIENCE_PREFIXES:
            if part.startswith(prefix) or prefix in part:
                # Извлекаем аудиторию (может быть с номером)
                # Паттерн: префикс + номер (например, "У708", "К506", "А601")
                match = re.search(rf'({prefix}[^\s,;]*)', part, re.IGNORECASE)
                if match:
                    aud = match.group(1).strip()
                    if aud not in audiences:
                        audiences.append(aud)
                else:
                    if part not in audiences:
                        audiences.append(part)
                break
    
    # Если не нашли по префиксам, ищем паттерны типа "ауд. 708", "ауд 506"
    if not audiences:
        # Ищем "ауд" или "аудитория"
        aud_pattern = re.search(r'(?:ауд\.?|аудитория)\s*([А-ЯЁа-яё0-9\-\s]+)', text, re.IGNORECASE)
        if aud_pattern:
            audiences.append(aud_pattern.group(1).strip())
    
    return ', '.join(audiences) if audiences else ''

def extract_subject_name(text):
    """Извлекает название предмета из текста, убирая лишние символы"""
    if not text or not isinstance(text, str):
        return ''
    
    text = text.strip()
    
    # Убираем маркеры типа занятия
    text = re.sub(r'\(лек\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(пр\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(практика\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(л\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\(п\)', '', text, flags=re.IGNORECASE)
    
    # Убираем информацию о часах в скобках: (24 ч), (36 ч), (48 часов) и т.д.
    text = re.sub(r'\(\d+\s*(?:ч|час|часов|часа)?\)', '', text, flags=re.IGNORECASE)
    
    # СНАЧАЛА обрабатываем "//" - это разделитель числителя/знаменателя, может содержать разные дисциплины
    # Разделяем по "//" и обрабатываем каждую часть отдельно
    if '//' in text:
        parts_by_week = text.split('//')
        all_subject_parts = []
        
        for part in parts_by_week:
            part = part.strip()
            if not part:
                continue
            
            # Убираем маркеры типа занятия для этой части
            part = re.sub(r'\(лек\)', '', part, flags=re.IGNORECASE)
            part = re.sub(r'\(пр\)', '', part, flags=re.IGNORECASE)
            part = re.sub(r'\(практика\)', '', part, flags=re.IGNORECASE)
            part = re.sub(r'\(л\)', '', part, flags=re.IGNORECASE)
            part = re.sub(r'\(п\)', '', part, flags=re.IGNORECASE)
            
            # Убираем информацию о часах
            part = re.sub(r'\(\d+\s*(?:ч|час|часов|часа)?\)', '', part, flags=re.IGNORECASE)
            
            # Убираем "/" (одиночные)
            part = re.sub(r'\s*/\s*', ' ', part)
            
            # Разделяем по запятым и берем части, которые не являются аудиториями
            sub_parts = re.split(r'[,;]', part)
            for sub_part in sub_parts:
                sub_part = sub_part.strip()
                if not sub_part:
                    continue
                
                # Проверяем, не является ли это аудиторией
                # "ауд" без цифр - это не аудитория, это просто слово
                is_aud = False
                for prefix in AUDIENCE_PREFIXES:
                    # Проверяем, что префикс - это не просто первая буква слова
                    # Аудитория должна быть типа "У708", "К506", "А539" и т.д.
                    if sub_part.startswith(prefix):
                        # Если после префикса идет цифра или пробел+цифра - это аудитория
                        if len(sub_part) > len(prefix):
                            rest = sub_part[len(prefix):].strip()
                            if rest and (rest[0].isdigit() or rest.startswith(' ')):
                                is_aud = True
                                break
                    # Также проверяем, если префикс в середине/конце (для сложных префиксов)
                    elif prefix in sub_part and len(prefix) > 1:
                        # Для многосимвольных префиксов проверяем контекст
                        idx = sub_part.find(prefix)
                        if idx >= 0 and idx + len(prefix) < len(sub_part):
                            rest = sub_part[idx + len(prefix):].strip()
                            if rest and rest[0].isdigit():
                                is_aud = True
                                break
                
                # Проверяем паттерны аудитории с цифрами: "ауд. 708", "ауд 708"
                if not is_aud:
                    if re.search(r'ауд\.?\s*\d+', sub_part, re.IGNORECASE):
                        is_aud = True
                    # "ауд" без цифр - не аудитория
                    elif sub_part.lower().strip() == 'ауд':
                        is_aud = False
                
                if not is_aud and sub_part and sub_part not in all_subject_parts:
                    all_subject_parts.append(sub_part)
        
        subject = ', '.join(all_subject_parts) if all_subject_parts else ''
    else:
        # Если нет "//", обрабатываем как обычно
        # Убираем "/" (одиночные)
        text = re.sub(r'\s*/\s*', ' ', text)
        
        # Убираем информацию об аудитории
        # Разделяем по запятым и берем части, которые не являются аудиториями
        parts = re.split(r'[,;]', text)
        subject_parts = []
        
        for part in parts:
            part = part.strip()
            if not part:
                continue
            
            # Проверяем, не является ли это аудиторией
            is_aud = False
            for prefix in AUDIENCE_PREFIXES:
                if part.startswith(prefix) or prefix in part:
                    is_aud = True
                    break
            
            # Проверяем паттерны аудитории
            if not is_aud:
                if re.search(r'ауд\.?\s*\d+', part, re.IGNORECASE):
                    is_aud = True
            
            if not is_aud:
                subject_parts.append(part)
        
        # Если не нашли части без аудитории, берем первую часть
        if not subject_parts:
            subject_parts = [parts[0].strip()] if parts else [text]
        
        subject = ', '.join(subject_parts)
    
    # Убираем подгруппы в конце (типа "п/г 1", "п/г 2")
    subject = re.sub(r',?\s*п/г\s*\d+', '', subject, flags=re.IGNORECASE)
    subject = re.sub(r',?\s*подгруппа\s*\d+', '', subject, flags=re.IGNORECASE)
    
    # Убираем информацию об аудитории, которая могла остаться
    for prefix in AUDIENCE_PREFIXES:
        # Убираем паттерны типа "У902", "К506" и т.д. (в любом месте)
        subject = re.sub(rf',?\s*{re.escape(prefix)}\d+', '', subject, flags=re.IGNORECASE)
        subject = re.sub(rf',?\s*{re.escape(prefix)}\s*\d+', '', subject, flags=re.IGNORECASE)
        # Убираем паттерны типа ". У902", ". К506" (точка перед аудиторией)
        subject = re.sub(rf'\s*\.\s*{re.escape(prefix)}\d+', '', subject, flags=re.IGNORECASE)
        subject = re.sub(rf'\s*\.\s*{re.escape(prefix)}\s*\d+', '', subject, flags=re.IGNORECASE)
        # Убираем паттерны типа " . У902" (пробел, точка, пробел, аудитория)
        subject = re.sub(rf'\s+\.\s+{re.escape(prefix)}\d+', '', subject, flags=re.IGNORECASE)
        subject = re.sub(rf'\s+\.\s+{re.escape(prefix)}\s*\d+', '', subject, flags=re.IGNORECASE)
    
    # Убираем паттерны "ауд. 708", "ауд 506"
    subject = re.sub(r',?\s*ауд\.?\s*\d+', '', subject, flags=re.IGNORECASE)
    
    # Убираем оставшиеся "/" и "//" в любом месте
    subject = re.sub(r'\s*//+\s*', ' ', subject)
    subject = re.sub(r'\s*/\s*', ' ', subject)
    
    # Убираем точки, которые остались после удаления аудиторий
    subject = re.sub(r'\s*\.\s*$', '', subject)  # точка в конце
    subject = re.sub(r'\s*\.\s*\.', '.', subject)  # множественные точки
    subject = re.sub(r'\s+\.\s+', ' ', subject)  # точка с пробелами
    subject = re.sub(r'^\s*\.\s*', '', subject)  # точка в начале
    
    # Убираем лишние символы и очищаем
    # Убираем множественные пробелы
    subject = ' '.join(subject.split())
    # Убираем пробелы вокруг запятых
    subject = re.sub(r'\s*,\s*', ', ', subject)
    # Убираем запятые в начале и конце
    subject = subject.strip(',').strip()
    # Убираем "/" и "//" в начале и конце
    subject = re.sub(r'^[/\s]+', '', subject)
    subject = re.sub(r'[/\s]+$', '', subject)
    
    return subject.strip()

def find_header_row(ws):
    """Находит строку с заголовками (Дисциплина, Преподаватель и т.д.)"""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row = ws[row_idx]
        has_discipline = False
        has_teacher = False
        has_pair = False
        
        for cell in row:
            if cell.value:
                value = str(cell.value).lower()
                # Ищем различные варианты написания
                if 'дисциплина' in value or 'дисц' in value:
                    has_discipline = True
                if 'преподаватель' in value or 'препод' in value:
                    has_teacher = True
                if 'пара' in value:
                    has_pair = True
        
        if has_discipline and has_teacher:
            return row_idx
    
    return None

def find_column_indices(ws, header_row):
    """Находит индексы столбцов для дисциплины, преподавателя, пары и дня недели"""
    discipline_cols = []
    teacher_cols = []
    pair_col = None
    day_col = None
    
    row = ws[header_row]
    for col_idx, cell in enumerate(row, 1):
        if cell.value:
            value = str(cell.value).lower()
            if 'дисциплина' in value:
                discipline_cols.append(col_idx)
            if 'преподаватель' in value:
                teacher_cols.append(col_idx)
            if 'пара' in value:
                pair_col = col_idx
            if 'д/н' in value or 'день' in value and 'недели' in value:
                day_col = col_idx
    
    return discipline_cols, teacher_cols, pair_col, day_col

def find_day_blocks(ws, header_row):
    """Находит блоки дней недели в файле"""
    day_blocks = []
    
    # Ищем названия дней недели в строках перед заголовками
    # Обычно дни недели находятся в строках 6-8
    for row_idx in range(max(1, header_row - 5), header_row):
        row = ws[row_idx]
        found_days_in_row = []
        
        for col_idx, cell in enumerate(row, 1):
            if cell.value:
                value = str(cell.value).lower()
                for day in DAYS_OF_WEEK:
                    if day in value and col_idx not in [b['start_col'] for b in day_blocks]:
                        found_days_in_row.append((col_idx, day))
        
        # Если нашли дни недели в этой строке, создаем блоки
        for col_idx, day in found_days_in_row:
            # Ищем конец блока (следующий день или конец файла)
            block_end_col = None
            for next_col in range(col_idx + 1, min(ws.max_column + 1, col_idx + 10)):
                if next_col <= ws.max_column:
                    next_cell = ws.cell(row=row_idx, column=next_col)
                    if next_cell.value:
                        next_value = str(next_cell.value).lower()
                        if any(d in next_value for d in DAYS_OF_WEEK):
                            block_end_col = next_col
                            break
            
            if not block_end_col:
                # Если не нашли следующий день, берем конец файла или следующую позицию через несколько столбцов
                block_end_col = min(col_idx + 7, ws.max_column + 1)
            
            day_blocks.append({
                'day': day,
                'start_col': col_idx,
                'end_col': block_end_col,
                'header_row': header_row
            })
    
    # Если не нашли дни недели, создаем один блок для всего файла
    if not day_blocks:
        day_blocks.append({
            'day': 'неизвестно',
            'start_col': 1,
            'end_col': ws.max_column + 1,
            'header_row': header_row
        })
    
    return day_blocks

def get_merged_cell_value(ws, cell):
    """Получает значение merged ячейки (берет из верхней-левой ячейки merged range)"""
    if cell.value is not None:
        return cell.value
    
    # Проверяем, входит ли ячейка в merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Берем значение из верхней-левой ячейки merged range
            top_left_cell = ws[merged_range.min_row][merged_range.min_col - 1]
            return top_left_cell.value
    
    return None

def parse_day_of_week(day_value):
    """Парсит день недели из значения"""
    if not day_value:
        return None
    
    day_str = str(day_value).lower().strip()
    
    # Убираем точки и пробелы в конце для точного сравнения
    day_str_clean = day_str.rstrip('. ')
    
    # Маппинг дней недели (проверяем точное совпадение или начало строки)
    # Порядок важен: сначала длинные варианты, потом короткие
    if day_str_clean.startswith('понедельник') or day_str_clean in ['пн', '1']:
        return 'понедельник'
    elif day_str_clean.startswith('вторник') or day_str_clean in ['вт', '2']:
        return 'вторник'
    elif day_str_clean.startswith('среда') or day_str_clean in ['ср', '3']:
        return 'среда'
    elif day_str_clean.startswith('четверг') or day_str_clean in ['чт', '4']:
        return 'четверг'
    elif day_str_clean.startswith('пятница') or day_str_clean in ['пт', '5']:
        return 'пятница'
    elif day_str_clean.startswith('суббота') or day_str_clean in ['сб', '6']:
        return 'суббота'
    elif day_str_clean.startswith('воскресенье') or day_str_clean in ['вс', '7']:
        return 'воскресенье'
    
    return None

def extract_schedule_metadata(ws, start_col, end_col):
    """Извлекает метаданные расписания (группа, институт, курс) из блока"""
    group = ''
    institute = ''
    course = ''
    direction = ''
    
    # Ищем в строках 6-9 (где обычно находятся эти данные)
    for row_idx in range(6, min(10, ws.max_row + 1)):
        row = ws[row_idx]
        
        # Ищем метки и значения в этом блоке
        for col_idx in range(start_col, min(end_col, len(row) + 1)):
            if col_idx <= len(row):
                cell = row[col_idx - 1]
                if cell.value:
                    value = str(cell.value).strip()
                    value_lower = value.lower()
                    
                    # Проверяем соседние ячейки для контекста
                    # Ищем группу (обычно рядом с текстом "группа" или формат "606-51")
                    if re.match(r'^\d{3}-\d{2}', value):
                        # Проверяем, есть ли рядом слово "группа"
                        if col_idx < len(row):
                            next_cell = row[col_idx]
                            if next_cell.value and 'группа' in str(next_cell.value).lower():
                                group = value
                        else:
                            group = value
                    
                    # Ищем институт (обычно в строке 7, колонка 3)
                    if row_idx == 7:
                        # Проверяем, есть ли в этой строке метка "институт"
                        for check_col in range(start_col, min(end_col, len(row) + 1)):
                            if check_col <= len(row):
                                check_cell = row[check_col - 1]
                                if check_cell.value and 'институт' in str(check_cell.value).lower():
                                    # Институт обычно в следующей колонке
                                    if check_col < len(row):
                                        inst_cell = row[check_col]
                                        if inst_cell.value:
                                            institute = str(inst_cell.value).strip()
                                    break
                        # Если не нашли по метке, ищем по ключевым словам в колонке 3
                        if not institute and col_idx == start_col + 2:
                            if any(inst in value_lower for inst in ['политехнический', 'медицинский', 'гуманитарный', 'экономический']):
                                institute = value
                    
                    # Ищем курс (обычно в строке 7, колонка 5)
                    if row_idx == 7:
                        # Проверяем, есть ли в этой строке метка "курс"
                        for check_col in range(start_col, min(end_col, len(row) + 1)):
                            if check_col <= len(row):
                                check_cell = row[check_col - 1]
                                if check_cell.value:
                                    check_value = str(check_cell.value).lower()
                                    if 'курс' in check_value:
                                        # Курс обычно в следующей колонке
                                        if check_col < len(row):
                                            course_cell = row[check_col]
                                            if course_cell.value:
                                                course_val = str(course_cell.value).strip()
                                                if course_val.isdigit() and 1 <= int(course_val) <= 6:
                                                    course = course_val
                                        break
                        # Если не нашли по метке, ищем число 1-6 в колонке 5
                        if not course and col_idx == start_col + 4:
                            if value.isdigit() and 1 <= int(value) <= 6:
                                course = value
                    
                    # Ищем направление/специальность (обычно в строке 8, колонка 3)
                    if row_idx == 8:
                        # Проверяем колонку 3 (start_col + 2)
                        if col_idx == start_col + 2:
                            # Направление обычно начинается с цифр типа "09.03.01"
                            if re.match(r'^\d{2}\.\d{2}\.\d{2}', value):
                                direction = value
                        # Или ищем рядом с меткой "направление"
                        for check_col in range(start_col, min(end_col, len(row) + 1)):
                            if check_col <= len(row):
                                check_cell = row[check_col - 1]
                                if check_cell.value:
                                    check_value = str(check_cell.value).lower()
                                    if 'направление' in check_value:
                                        if check_col < len(row):
                                            dir_cell = row[check_col]
                                            if dir_cell.value:
                                                direction = str(dir_cell.value).strip()
                                        break
    
    return {
        'group': group,
        'institute': institute,
        'course': course,
        'direction': direction
    }

def find_schedule_tables(ws, header_row):
    """Находит все таблицы расписаний на странице по количеству вхождений слова 'дисциплина'"""
    tables = []
    
    header_row_data = ws[header_row]
    discipline_cols = []
    teacher_cols = []
    pair_cols = []
    day_cols = []
    
    # Ищем все вхождения "дисциплина" и связанные колонки
    for col_idx, cell in enumerate(header_row_data, 1):
        if cell.value:
            value = str(cell.value).lower()
            if 'дисциплина' in value:
                discipline_cols.append(col_idx)
            if 'преподаватель' in value:
                teacher_cols.append(col_idx)
            if 'пара' in value:
                pair_cols.append(col_idx)
            if 'д/н' in value or ('день' in value and 'недели' in value):
                day_cols.append(col_idx)
    
    # Если нет дисциплин, возвращаем пустой список
    if not discipline_cols:
        return tables
    
    # Для каждой колонки "дисциплина" создаем отдельную таблицу
    for disc_col in discipline_cols:
        # Находим начало таблицы (ищем колонку "д/н", "№/п" или "пара" слева от дисциплины)
        start_col = disc_col
        # Ищем ближайшую колонку с заголовком слева от дисциплины
        for col_idx in range(disc_col - 1, 0, -1):
            if col_idx <= len(header_row_data):
                cell = header_row_data[col_idx - 1]
                if cell.value:
                    value = str(cell.value).lower()
                    if 'д/н' in value or '№/п' in value or 'пара' in value:
                        start_col = col_idx
                        break
        # Если не нашли, берем предыдущую дисциплину как начало (если есть)
        if start_col == disc_col and disc_col > 1:
            prev_disc = None
            for prev_disc_col in discipline_cols:
                if prev_disc_col < disc_col:
                    prev_disc = prev_disc_col
            if prev_disc:
                # Начало - это колонка после предыдущей дисциплины
                start_col = prev_disc + 1
            else:
                # Если это первая дисциплина, начинаем с 1
                start_col = 1
        
        # Находим конец таблицы (колонка "преподаватель" справа от дисциплины)
        end_col = disc_col + 1
        for teacher_col in teacher_cols:
            if teacher_col > disc_col:
                end_col = teacher_col + 1
                break
        
        # Если не нашли преподавателя справа, ищем до следующей дисциплины или до конца
        if end_col == disc_col + 1:
            # Ищем следующую дисциплину
            next_disc = None
            for next_disc_col in discipline_cols:
                if next_disc_col > disc_col:
                    next_disc = next_disc_col
                    break
            if next_disc:
                end_col = next_disc
            else:
                # Ищем преподавателя в любом месте после дисциплины
                for teacher_col in teacher_cols:
                    if teacher_col > disc_col:
                        end_col = teacher_col + 1
                        break
                if end_col == disc_col + 1:
                    end_col = ws.max_column + 1
        
        # Находим колонки для этой таблицы
        # КРИТИЧНО: колонка "д/н" может быть ДО начала таблицы (в первой колонке),
        # поэтому ищем её во всем заголовке, а не только в пределах таблицы
        table_pair_col = None
        table_day_col = None
        table_teacher_col = None
        
        # Сначала ищем колонку "д/н" во всем заголовке (может быть в первой колонке)
        for col_idx in range(1, len(header_row_data) + 1):
            if col_idx <= len(header_row_data):
                cell = header_row_data[col_idx - 1]
                if cell.value:
                    value = str(cell.value).lower()
                    if ('д/н' in value or ('день' in value and 'недели' in value)) and not table_day_col:
                        table_day_col = col_idx
                        break  # Нашли, выходим
        
        # Ищем колонки "пара" и "преподаватель" в пределах таблицы
        for col_idx in range(start_col, min(end_col, len(header_row_data) + 1)):
            if col_idx <= len(header_row_data):
                cell = header_row_data[col_idx - 1]
                if cell.value:
                    value = str(cell.value).lower()
                    if 'пара' in value and not table_pair_col:
                        table_pair_col = col_idx
                    if 'преподаватель' in value and not table_teacher_col:
                        table_teacher_col = col_idx
        
        tables.append({
            'start_col': start_col,
            'end_col': end_col,
            'discipline_col': disc_col,
            'teacher_col': table_teacher_col,
            'pair_col': table_pair_col,
            'day_col': table_day_col,
            'header_row': header_row
        })
    
    return tables

def parse_excel_sheet(ws, teacher_name_mapping, course_from_sheet=None):
    """Парсит один лист Excel файла с расписанием (может быть несколько таблиц на странице)"""
    results = []
    
    # Находим строку с заголовками
    header_row = find_header_row(ws)
    if not header_row:
        return results
    
    # Находим все таблицы расписаний на странице (по количеству "дисциплина")
    schedule_tables = find_schedule_tables(ws, header_row)
    
    if not schedule_tables:
        return results
    
    # Извлекаем метаданные один раз для всего листа
    metadata = extract_schedule_metadata(ws, 1, ws.max_column + 1)
    if not metadata['course'] and course_from_sheet:
        metadata['course'] = course_from_sheet
    
    # Обрабатываем каждую таблицу
    for table in schedule_tables:
        start_col = table['start_col']
        end_col = table['end_col']
        disc_col = table['discipline_col']
        teacher_col = table['teacher_col']
        pair_col = table['pair_col']
        day_col = table['day_col']
        
        # КРИТИЧНО: Сбрасываем current_day при начале каждой новой таблицы
        current_day = None
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            
            # Пропускаем полностью пустые строки в пределах таблицы
            has_data = False
            for col_idx in range(start_col, min(end_col, len(row) + 1)):
                if col_idx <= len(row) and row[col_idx - 1].value:
                    has_data = True
                    break
            if not has_data:
                continue
            
            # КРИТИЧНО: Получаем день недели СРАЗУ и обновляем current_day,
            # даже если строка будет пропущена из-за пустой дисциплины
            day_of_week = None
            
            # Сначала проверяем колонку "д/н" (с учетом merged ячеек)
            if day_col and day_col <= len(row):
                day_cell = row[day_col - 1]
                # Проверяем merged ячейки
                day_value = get_merged_cell_value(ws, day_cell)
                if day_value:
                    day_of_week = parse_day_of_week(day_value)
                    if day_of_week:
                        # КРИТИЧНО: Обновляем current_day немедленно, даже если строка будет пропущена
                        current_day = day_of_week
            
            # Если не нашли, проверяем первую колонку (колонка 1, где обычно находится день)
            # КРИТИЧНО: колонка "д/н" может быть в первой колонке, которая находится ДО start_col
            if not day_of_week and len(row) > 0:
                first_cell = row[0]  # Первая колонка листа (не таблицы)
                first_value = get_merged_cell_value(ws, first_cell)
                if first_value:
                    parsed_day = parse_day_of_week(first_value)
                    if parsed_day:
                        day_of_week = parsed_day
                        # КРИТИЧНО: Обновляем current_day немедленно
                        current_day = parsed_day
            
            # Если день недели не найден в этой строке, используем день из предыдущей строки
            if not day_of_week:
                day_of_week = current_day
            
            # Если день недели все еще не найден, пропускаем строку
            if not day_of_week:
                continue
            
            # Получаем номер пары (ОБЯЗАТЕЛЬНО)
            pair_numbers = []
            if pair_col and pair_col <= len(row):
                pair_cell = row[pair_col - 1]
                if pair_cell.value:
                    pair_numbers = parse_pair_number(pair_cell.value)
            
            # Если не нашли в колонке "пара", проверяем вторую колонку таблицы
            if not pair_numbers and start_col + 1 <= len(row):
                second_cell = row[start_col]
                if second_cell.value:
                    pair_numbers = parse_pair_number(second_cell.value)
            
            # Если номер пары не найден, пропускаем строку
            if not pair_numbers:
                continue
            
            # Получаем дисциплину (ОБЯЗАТЕЛЬНО) - сырой текст всех ячеек от "дисциплина" до "преподаватель"
            # Находим конец блока дисциплин (до преподавателя, не включая его)
            discipline_end_col = teacher_col if teacher_col and teacher_col > disc_col else end_col
            
            # Собираем сырой текст всех ячеек от дисциплины до преподавателя
            discipline_parts = []
            for col_idx in range(disc_col, min(discipline_end_col, len(row) + 1)):
                if col_idx <= len(row):
                    cell = row[col_idx - 1]
                    if cell and cell.value:
                        disc_text = str(cell.value).strip()
                        if disc_text and disc_text.lower() not in ['none', 'null', '']:
                            discipline_parts.append(disc_text)
            
            # Если нет дисциплин (все ячейки пустые), пропускаем строку
            if not discipline_parts:
                continue
            
            # Объединяем все ячейки дисциплины в один сырой текст (через пробел)
            # Это обязательное поле - сырой текст блока дисциплины
            raw_discipline = ' '.join(discipline_parts).strip()
            
            # Если после объединения текст пустой, пропускаем
            if not raw_discipline:
                continue
            
            # Опционально: извлекаем дополнительные данные (если не ломает обязательное)
            # Преподаватель
            teacher_fio = ''
            if teacher_col and teacher_col <= len(row):
                teacher_cell = row[teacher_col - 1]
                if teacher_cell.value:
                    teacher_text = str(teacher_cell.value).strip()
                    if teacher_text:
                        try:
                            normalized_short_fio = normalize_short_fio(teacher_text)
                            teacher_fio = teacher_name_mapping.get(normalized_short_fio, teacher_text)
                        except:
                            teacher_fio = teacher_text
            
            # Опционально: аудитория, тип занятия, неделя
            audience = extract_audience(raw_discipline)
            lecture_type = parse_lecture_type(raw_discipline)
            week_types = parse_week_type(raw_discipline)
            is_remote = 'ЭОиДОТ' in raw_discipline.upper() or 'эоидот' in raw_discipline.lower()
            
            if not week_types:
                week_types = ['обе недели']
            
            # Создаем записи для каждой пары (если пара 1-2, то две записи)
            for pair_num in pair_numbers:
                # Если есть разделение по неделям, создаем отдельные записи
                for week_type in week_types:
                    result_entry = {
                        # ОБЯЗАТЕЛЬНЫЕ поля (100%)
                        'day_of_week': day_of_week,
                        'pair_number': pair_num,
                        'subject_name': raw_discipline,  # Сырой текст, как есть
                        
                        # Опциональные поля
                        'teacher': teacher_fio,
                        'audience': audience,
                        'lecture_type': lecture_type,
                        'week_type': week_type,
                        'is_remote': is_remote,
                        'is_external': False,
                        'department': '',
                        'group': metadata['group'],
                        'institute': metadata['institute'],
                        'course': metadata['course'],
                        'direction': metadata['direction']
                    }
                    results.append(result_entry)
    
    return results

def parse_excel_file(file_path, teacher_name_mapping):
    """Парсит Excel файл с расписанием (все листы)"""
    wb = load_workbook(file_path, data_only=True)
    
    all_results = []
    
    # Обрабатываем все листы (каждый лист - это курс)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Извлекаем номер курса из названия листа
        course_from_sheet = None
        course_match = re.search(r'(\d+)\s*курс', sheet_name, re.IGNORECASE)
        if course_match:
            course_from_sheet = course_match.group(1)
        
        sheet_results = parse_excel_sheet(ws, teacher_name_mapping, course_from_sheet)
        all_results.extend(sheet_results)
    
    return all_results

def save_results_to_csv(results, output_file):
    """Сохраняет результаты в CSV файл со всеми доступными полями"""
    import csv
    
    if not results:
        return
    
    # Собираем все уникальные поля из всех результатов
    all_fields = set()
    for result in results:
        all_fields.update(result.keys())
    
    # Определяем порядок полей (приоритетные сначала)
    priority_fields = [
        'day_of_week', 'pair_number', 'subject_name', 
        'teacher', 'fio', 'audience', 'lecture_type', 
        'week_type', 'week', 'group', 'group_name',
        'subgroup', 'num_subgroups', 'is_external', 
        'is_remote', 'department', 'institute', 
        'course', 'direction'
    ]
    
    # Формируем финальный список полей
    fieldnames = []
    # Сначала добавляем приоритетные поля, которые есть в данных
    for field in priority_fields:
        if field in all_fields:
            fieldnames.append(field)
            all_fields.discard(field)
    
    # Затем добавляем остальные поля в алфавитном порядке
    fieldnames.extend(sorted(all_fields))
    
    # Убеждаемся, что обязательные поля есть
    required_fields = ['day_of_week', 'pair_number', 'subject_name']
    for field in required_fields:
        if field not in fieldnames:
            fieldnames.insert(0, field)
    
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in results:
            # Создаем строку со всеми полями
            row = {}
            for field in fieldnames:
                value = result.get(field, '')
                # Преобразуем None в пустую строку
                if value is None:
                    value = ''
                # Преобразуем булевы значения
                elif isinstance(value, bool):
                    value = 'True' if value else 'False'
                row[field] = value
            writer.writerow(row)
    
    print(f"CSV сохранен с полями: {fieldnames}")

def save_results_to_excel(results, output_file):
    """Сохраняет результаты в Excel файл в формате, похожем на timetable_processed.csv"""
    if not results:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    # Заголовки (точно как в CSV)
    headers = [
        'fio', 'pair_number', 'day_of_week', 'group', 'audience', 'department',
        'week', 'subgroup', 'num_subgroups', 'is_external', 'is_remote', 'subject_name'
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Записываем данные
    for row_num, result in enumerate(results, 2):
        # Маппинг полей
        values = {
            'fio': result.get('teacher', ''),
            'pair_number': result.get('pair_number', ''),
            'day_of_week': result.get('day_of_week', ''),
            'group': result.get('group', ''),
            'audience': result.get('audience', ''),
            'department': result.get('department', ''),  # Может быть пустым, если не извлекали
            'week': result.get('week_type', ''),
            'subgroup': result.get('subgroup', ''),
            'num_subgroups': result.get('num_subgroups', ''),  # Может быть пустым
            'is_external': result.get('is_external', False),  # Может быть пустым
            'is_remote': result.get('is_remote', False),
            'subject_name': result.get('subject_name', '')
        }
        
        for col_num, header in enumerate(headers, 1):
            value = values.get(header, '')
            # Преобразуем булевы значения в строки как в CSV
            if isinstance(value, bool):
                value = 'True' if value else 'False'
            # Преобразуем числа в строки
            elif isinstance(value, (int, float)):
                value = str(value)
            # Пустые значения оставляем пустыми
            elif value == '' or value is None:
                value = ''
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Замораживаем первую строку
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)

def main():
    import os
    
    # Создаем папки
    os.makedirs('input/timetable', exist_ok=True)
    os.makedirs('output/timetable', exist_ok=True)
    
    # Загружаем маппинг ФИО преподавателей
    teacher_name_mapping = load_teacher_names()
    if teacher_name_mapping:
        print(f"Загружено {len(teacher_name_mapping)} полных ФИО преподавателей")
    
    # Ищем все Excel файлы в папке input/timetable
    excel_files = glob.glob("input/timetable/*.xlsx")
    
    # Фильтруем временные файлы Excel (начинающиеся с ~$)
    excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
    
    if not excel_files:
        print("Не найдены Excel файлы в папке input/timetable/")
        return
    
    print(f"Найдено файлов: {len(excel_files)}")
    
    # Собираем все результаты из всех файлов
    all_results = []
    
    # Обрабатываем каждый файл
    for file_path in excel_files:
        filename = os.path.basename(file_path)
        print(f"\nОбрабатываем файл: {filename}")
        
        try:
            results = parse_excel_file(file_path, teacher_name_mapping)
            print(f"Обработано записей: {len(results)}")
            all_results.extend(results)
        except Exception as e:
            print(f"Ошибка при обработке файла {filename}: {e}")
            import traceback
            traceback.print_exc()
    
    # Сохраняем все результаты в один файл
    if all_results:
        # Сохраняем в CSV с обязательными полями
        csv_output_file = os.path.join('output/timetable', 'timetable_processed.csv')
        save_results_to_csv(all_results, csv_output_file)
        print(f"CSV сохранен в: {csv_output_file}")
        
        # Сохраняем в Excel с полными данными
        excel_output_file = os.path.join('output/timetable', 'timetable_processed.xlsx')
        save_results_to_excel(all_results, excel_output_file)
        print(f"\nВсего обработано записей: {len(all_results)}")
        print(f"Excel сохранен в: {excel_output_file}")
    else:
        print("Не удалось извлечь данные из файлов")

if __name__ == '__main__':
    main()
