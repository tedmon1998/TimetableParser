import csv
import json
import re
import os
from collections import defaultdict

# Маппинг дней недели
DAYS_MAPPING = {
    3: 'понедельник',
    5: 'вторник',
    7: 'среда',
    9: 'четверг',
    11: 'пятница',
    13: 'суббота'
}

# Маппинг столбцов аудиторий
AUDIENCE_COLUMNS = {
    3: 4,   # понедельник -> аудитория
    5: 6,   # вторник -> аудитория.
    7: 8,   # среда -> аудитория..
    9: 10,  # четверг -> аудитория_
    11: 12, # пятница -> аудитория_
    13: 14  # суббота -> аудитория…
}

# Маппинг недель на русский
WEEK_MAPPING = {
    'both': 'обе недели',
    'numerator': 'числитель',
    'denominator': 'знаменатель'
}

def parse_group_string(group_str):
    """
    Парсит строку с группами и возвращает список словарей с информацией о каждой группе
    Примеры:
    - "601-31" -> [{group: "601-31", week: "both", subgroup: None}]
    - "605-41/" -> [{group: "605-41", week: "numerator", subgroup: None}]
    - "/601-51м" -> [{group: "601-51м", week: "denominator", subgroup: None}]
    - "601-21а" -> [{group: "601-21", week: "both", subgroup: "а"}]
    - "601-51аб" -> [{group: "601-51", week: "both", subgroup: "а"}, {group: "601-51", week: "both", subgroup: "б"}]
    - "607-51,607-52/" -> [{group: "607-51", week: "numerator", subgroup: None}, {group: "607-52", week: "numerator", subgroup: None}]
    """
    if not group_str or group_str.strip() == '':
        return []
    
    groups = []
    group_str = group_str.strip()
    
    # Определяем числитель/знаменатель для всей строки
    week_type = "both"  # по умолчанию обе недели
    if group_str.endswith('/'):
        week_type = "numerator"  # числитель
        group_str = group_str[:-1]
    elif group_str.startswith('/'):
        week_type = "denominator"  # знаменатель
        group_str = group_str[1:]
    
    # Разделяем по запятым, но также учитываем слэши внутри (которые могут разделять группы)
    # Сначала разделяем по запятым
    parts = [p.strip() for p in group_str.split(',')]
    group_list = []
    
    for part in parts:
        # Если в части есть слэш, который разделяет группы (не в начале и не в конце)
        # Например: "607-51/607-52" -> ["607-51", "607-52"]
        if '/' in part and not part.startswith('/') and not part.endswith('/'):
            # Разделяем по слэшу
            sub_parts = [p.strip() for p in part.split('/')]
            group_list.extend(sub_parts)
        else:
            group_list.append(part)
    
    # Буквы, которые являются подгруппами (исключаем "м" - магистратура)
    subgroup_letters = set('абвгдежзиклнопрстуфхцчшщэюя')
    
    for group in group_list:
        if not group:
            continue
        
        # Извлекаем подгруппы (а, б, в, г и т.д., но не "м")
        # Ищем паттерн типа "601-51аб" или "601-21а" в конце строки
        # Подгруппы идут после номера группы без дефиса
        base_group = group
        subgroups = []
        
        # Ищем буквы в конце, которые могут быть подгруппами
        # Паттерн: номер группы, затем возможные буквы подгрупп
        # Например: "601-21а", "601-51аб", но не "601-51м" (м - магистратура)
        match = re.search(r'^(.+?)([абвгдежзиклнопрстуфхцчшщэюя]+)$', group, re.IGNORECASE)
        
        if match:
            potential_base = match.group(1)
            potential_subgroups = match.group(2).lower()
            
            # Проверяем, что все буквы являются подгруппами (не "м")
            valid_subgroups = [char for char in potential_subgroups if char in subgroup_letters]
            
            if valid_subgroups and len(valid_subgroups) == len(potential_subgroups):
                # Все буквы - подгруппы
                base_group = potential_base
                subgroups = valid_subgroups
            else:
                # Есть буквы, которые не являются подгруппами (например, "м")
                # Значит это часть названия группы
                base_group = group
                subgroups = []
        
        if subgroups:
            # Если есть подгруппы, создаем запись для каждой
            for subgroup in subgroups:
                groups.append({
                    'group': base_group,
                    'week': week_type,
                    'subgroup': subgroup
                })
        else:
            # Если подгрупп нет
            groups.append({
                'group': base_group,
                'week': week_type,
                'subgroup': None
            })
    
    return groups

def count_subgroups(groups_list):
    """Подсчитывает количество уникальных подгрупп"""
    subgroups = set()
    for group_info in groups_list:
        if group_info['subgroup']:
            subgroups.add(group_info['subgroup'])
    return len(subgroups) if subgroups else 0

def normalize_short_fio(short_fio):
    """Нормализует короткое ФИО для сопоставления"""
    if not short_fio:
        return short_fio
    
    # Убираем лишние пробелы
    short_fio = ' '.join(short_fio.split())
    
    # Приводим к стандартному формату: "Фамилия И.О."
    # Убираем пробелы между инициалами и точками
    short_fio = re.sub(r'([А-ЯЁ])\.\s*([А-ЯЁ])\.', r'\1.\2.', short_fio)
    short_fio = re.sub(r'([А-ЯЁ])\s+([А-ЯЁ])\.', r'\1.\2.', short_fio)
    
    return short_fio

def load_teacher_names(teacher_file='info/teacher_all.json'):
    """Загружает полные ФИО преподавателей из JSON файла и создает маппинг"""
    if not os.path.exists(teacher_file):
        print(f"Файл {teacher_file} не найден. Будет использоваться короткое ФИО.")
        return {}
    
    try:
        with open(teacher_file, 'r', encoding='utf-8') as f:
            teachers = json.load(f)
    except Exception as e:
        print(f"Ошибка при загрузке {teacher_file}: {e}. Будет использоваться короткое ФИО.")
        return {}
    
    # Создаем маппинг короткого ФИО на полное
    name_mapping = {}
    
    for teacher in teachers:
        full_fio = teacher.get('fio', '').strip()
        if not full_fio:
            continue
        
        # Парсим полное ФИО: "Галкин Владимир Александрович"
        parts = full_fio.split()
        if len(parts) >= 3:
            last_name = parts[0]
            first_name = parts[1]
            middle_name = parts[2]
            
            # Создаем различные варианты короткого ФИО
            variants = [
                f"{last_name} {first_name[0]}.{middle_name[0]}.",  # "Галкин В.А."
                f"{last_name} {first_name[0]}. {middle_name[0]}.",  # "Галкин В. А."
                f"{last_name} {first_name[0]}.{middle_name[0]}",    # "Галкин В.А" (без последней точки)
            ]
            
            for variant in variants:
                normalized = normalize_short_fio(variant)
                name_mapping[normalized] = full_fio
    
    return name_mapping

def process_csv_file(input_file, teacher_name_mapping=None):
    """Обрабатывает CSV файл и создает структурированные данные"""
    if teacher_name_mapping is None:
        teacher_name_mapping = {}
    
    results = []
    external_teachers = {}  # Словарь для хранения информации о внешних преподавателях
    missing_teachers = set()  # Множество преподавателей без полного ФИО
    
    # Сначала проходим по файлу и собираем информацию о внешних преподавателях
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)  # Пропускаем заголовок
        
        for row in reader:
            if len(row) < 16:
                continue
                
            teacher_fio = row[0].strip() if row[0] else ''
            if not teacher_fio or teacher_fio == 'вакансия':
                continue
            
            # Проверяем, является ли преподаватель внешним
            is_external = 'внешний' in (row[16].lower() if len(row) > 16 and row[16] else '')
            if is_external:
                external_teachers[teacher_fio] = True
    
    # Теперь обрабатываем файл и создаем записи
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)  # Пропускаем заголовок
        
        for row in reader:
            if len(row) < 16:
                continue
                
            teacher_fio = row[0].strip() if row[0] else ''
            department = row[1].strip() if row[1] else ''
            pair_number = row[2].strip() if row[2] else ''
            is_external = external_teachers.get(teacher_fio, False)
            
            if not teacher_fio or teacher_fio == 'вакансия':
                continue
            
            # Обрабатываем каждый день недели
            for day_col, day_name in DAYS_MAPPING.items():
                if day_col >= len(row):
                    continue
                    
                groups_str = row[day_col].strip() if row[day_col] else ''
                audience_col = AUDIENCE_COLUMNS[day_col]
                audience = row[audience_col].strip() if audience_col < len(row) and row[audience_col] else ''
                
                if not groups_str:
                    continue
                
                # Определяем дистанционность
                is_remote = audience.lower() == 'эоидот'
                
                # Парсим группы
                groups_list = parse_group_string(groups_str)
                
                if not groups_list:
                    continue
                
                # Подсчитываем количество подгрупп
                num_subgroups = count_subgroups(groups_list)
                
                # Создаем запись для каждой группы
                for group_info in groups_list:
                    # Получаем полное ФИО из маппинга, если доступно
                    normalized_short_fio = normalize_short_fio(teacher_fio)
                    full_fio = teacher_name_mapping.get(normalized_short_fio, teacher_fio)
                    
                    # Если полное ФИО не найдено, добавляем в список ошибок
                    if full_fio == teacher_fio and teacher_name_mapping:
                        missing_teachers.add(teacher_fio)
                    
                    # Преобразуем week в русский
                    week_ru = WEEK_MAPPING.get(group_info['week'], group_info['week'])
                    
                    result_entry = {
                        'fio': full_fio,
                        'pair_number': pair_number,
                        'day_of_week': day_name,
                        'group': group_info['group'],  # Номер группы
                        'audience': audience,
                        'department': department,
                        'week': week_ru,  # Теперь на русском
                        'subgroup': group_info['subgroup'] if group_info['subgroup'] else '',
                        'num_subgroups': num_subgroups,
                        'is_external': is_external,
                        'is_remote': is_remote,
                        'subject_name': ''  # Пока пустое, как просили
                    }
                    results.append(result_entry)
    
    return results, missing_teachers

def save_to_csv(data, output_file):
    """Сохраняет данные в CSV файл с UTF-8 BOM для корректного отображения в Excel"""
    if not data:
        return
    
    fieldnames = [
        'fio', 'pair_number', 'day_of_week', 'group', 'audience', 'department',
        'week', 'subgroup', 'num_subgroups', 'is_external', 'is_remote', 'subject_name'
    ]
    
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

def save_to_json(data, output_file):
    """Сохраняет данные в JSON файл"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_to_excel(data, output_file):
    """Сохраняет данные в Excel файл"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("Библиотека openpyxl не установлена. Установите её командой: pip install openpyxl")
        return
    
    if not data:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    # Заголовки
    headers = [
        'ФИО', 'Номер пары', 'День недели', 'Группа', 'Аудитория', 'Кафедра',
        'Неделя', 'Подгруппа', 'Кол-во подгрупп', 'Внешний', 'Дистанционно', 'Название предмета'
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Записываем данные
    fieldnames = [
        'fio', 'pair_number', 'day_of_week', 'group', 'audience', 'department',
        'week', 'subgroup', 'num_subgroups', 'is_external', 'is_remote', 'subject_name'
    ]
    
    for row_num, row_data in enumerate(data, 2):
        for col_num, field in enumerate(fieldnames, 1):
            value = row_data.get(field, '')
            # Преобразуем булевы значения в текст
            if isinstance(value, bool):
                value = 'Да' if value else 'Нет'
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Замораживаем первую строку
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)

def save_missing_teachers(missing_teachers, output_file='missing_teachers.csv'):
    """Сохраняет список преподавателей без полного ФИО в CSV файл с UTF-8 BOM"""
    if not missing_teachers:
        return
    
    # Сортируем для удобства
    sorted_teachers = sorted(missing_teachers)
    
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['short_fio'])  # Заголовок
        for teacher in sorted_teachers:
            writer.writerow([teacher])
    
    print(f"Сохранено {len(missing_teachers)} преподавателей без полного ФИО в {output_file}")

def main():
    # Ищем CSV файл с расписанием
    import glob
    import os
    
    # Создаем папки, если их нет
    os.makedirs('input', exist_ok=True)
    os.makedirs('output', exist_ok=True)
    
    # Ищем файлы, начинающиеся с "Zanyatost prepodavateley" в папке input
    csv_files = glob.glob("input/Zanyatost prepodavateley*.csv")
    
    if not csv_files:
        print("Не найден CSV файл с расписанием. Ожидается файл вида 'input/Zanyatost prepodavateley*.csv'")
        return
    
    input_file = csv_files[0]
    print(f"Обрабатываем файл: {input_file}")
    
    # Загружаем маппинг ФИО преподавателей
    teacher_name_mapping = load_teacher_names()
    if teacher_name_mapping:
        print(f"Загружено {len(teacher_name_mapping)} полных ФИО преподавателей")
    
    # Обрабатываем файл
    results, missing_teachers = process_csv_file(input_file, teacher_name_mapping)
    
    print(f"Обработано записей: {len(results)}")
    
    # Сохраняем результаты в папку output
    csv_output = 'output/timetable_processed.csv'
    json_output = 'output/timetable_processed.json'
    excel_output = 'output/timetable_processed.xlsx'
    
    save_to_csv(results, csv_output)
    print(f"Данные сохранены в CSV: {csv_output}")
    
    save_to_json(results, json_output)
    print(f"Данные сохранены в JSON: {json_output}")
    
    save_to_excel(results, excel_output)
    print(f"Данные сохранены в Excel: {excel_output}")
    
    # Сохраняем преподавателей без полного ФИО
    if missing_teachers:
        save_missing_teachers(missing_teachers, 'output/missing_teachers.csv')

if __name__ == '__main__':
    main()
